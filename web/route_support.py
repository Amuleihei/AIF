# 路由共享支持层：集中存放常量、工具函数与跨模块依赖
# 中文注释：route_modules 统一从这里取依赖，避免反向依赖 web.routes

from datetime import datetime
from io import BytesIO
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from flask import request, render_template_string, jsonify, redirect, url_for, flash, send_file
import logging
from flask_login import login_required, current_user

from aif import dispatch
from web.models import (
    Session,
    AdminAuditLog,
    FlowMetric,
    TrayBatch,
    ProductBatch,
    LogEntry,
    LogEntryMeta,
    LogEntryDetail,
    LogDriverProfile,
    LogPricingProfile,
    LogPricingRule,
    LogEntrySettlement,
    LogConsumption,
    SawRecord,
    SawMachineRecord,
    SawMachineLogDetail,
    DipRecord,
    SortRecord,
    ByproductRecord,
    generate_tray_batch_number,
    generate_product_batch_number,
)
from web.i18n import LANGUAGES  # 中文注释：从 i18n 聚合入口读取翻译
from web.utils import (
    get_lang,
    get_stock_data_with_lang,
    get_system_health_snapshot,
    update_flow_data,
    update_kiln_status,
)
from web.templates import HTML_TEMPLATE
from web.i18n.route_texts import ROUTE_TEXTS  # 中文注释：路由提示文案已拆分
from web.tray_parser import parse_sorted_kiln_trays, flatten_to_tray_items, count_total_trays, summarize_specs
from web.data_store import (
    get_flow_data,
    save_flow_data,
    get_kilns_data,
    save_kilns_data,
    get_log_stock_total,
    set_log_stock_total,
    add_log_stock,
    upsert_inventory_product,
    list_inventory_products,
    delete_inventory_product,
    get_shipping_data,
    save_shipping_data,
    get_inventory_products_by_ids,
    update_inventory_product_status,
)


BASE = Path(__file__).resolve().parent.parent
BARK_PRICE_PER_M3_KS = 31765.0
WASTE_SEGMENT_PRICE_PER_BAG_KS = 4000
SECONDARY_SPEC_PCS = {
    "220x81x21": 2160,
    "270x81x21": 1728,
    "370x81x21": 1248,
    "930x81x21": 517,
    "950x81x21": 517,
    "970x81x21": 517,
    "270x68x21": 2160,
    "370x68x21": 1536,
    "930x68x21": 654,
    "950x68x21": 654,
    "270x58x21": 2304,
    "370x58x21": 1632,
    "930x58x21": 705,
    "950x58x21": 705,
    "370x44x21": 2208,
    "950x44x21": 940,
    "270x44x21": 3024,
}
SECONDARY_SPEC_EXTRA_PCS = {
    "950x81x21": [528],
    "950x68x21": [658],
}
SECONDARY_RULE_PREFIX = "secondary_rule:"

def _lang_code():
    lang = get_lang()
    return lang if lang in ROUTE_TEXTS else "zh"


def _t(key):
    return ROUTE_TEXTS[_lang_code()].get(key, ROUTE_TEXTS["zh"].get(key, key))


def sync_raw_inventory(delta_mt: float):
    add_log_stock(delta_mt)


def _to_float(raw, default=0.0):
    try:
        if raw in (None, ""):
            return float(default)
        return float(raw)
    except Exception:
        return float(default)


def _to_int(raw, default=0):
    try:
        if raw in (None, ""):
            return int(default)
        return int(float(raw))
    except Exception:
        return int(default)


def audit_admin_action(action: str, target: str = "", detail: str = ""):
    """记录管理员审计日志。失败时静默，避免影响主业务。"""
    session = None
    try:
        session = Session()
        row = AdminAuditLog(
            operator=str(getattr(current_user, "username", "") or ""),
            action=str(action or ""),
            target=str(target or ""),
            detail=str(detail or ""),
        )
        session.add(row)
        session.commit()
    except Exception:
        pass
    finally:
        try:
            if session is not None:
                session.close()
        except Exception:
            pass


def _read_flow_data():
    return get_flow_data()


def _save_flow_data(flow: dict):
    save_flow_data(flow if isinstance(flow, dict) else {})


def _set_raw_log_stock(value: float):
    set_log_stock_total(value)


def _load_kilns_data():
    return get_kilns_data()


def _save_kilns_data(kilns: dict):
    save_kilns_data(kilns if isinstance(kilns, dict) else {})


def _parse_kiln_trays_input(raw: str, known_trays: dict, allow_plain_count: bool = False):
    tray_list = []
    unload_count = 0

    tokens = [t.strip() for t in (raw or "").replace("，", ",").split(",") if t.strip()]
    for token in tokens:
        if allow_plain_count and token.isdigit():
            count = _to_int(token, 0)
            if count <= 0:
                raise ValueError(f"bad tray count: {token}")
            tray_list.append({"id": f"TEMP-{len(tray_list)+1}", "spec": "", "count": count, "volume": count * 0.1, "batch_number": ""})
            unload_count += count
            continue
        if "#" in token:
            parts = token.split("#")
            if len(parts) < 3:
                raise ValueError(f"bad kiln token: {token}")
            tray_id = parts[0].strip()
            spec = parts[1].strip()
            count = _to_int(parts[2], 0)
            if count <= 0:
                raise ValueError(f"bad tray count: {token}")
            tray_list.append({"id": tray_id, "spec": spec, "count": count, "volume": count * 0.1, "batch_number": ""})
            unload_count += count
            continue

        tray_id = token
        tray_meta = known_trays.get(tray_id, {})
        spec = summarize_specs(tray_meta.get("specs", []))
        tray_list.append({"id": tray_id, "spec": spec, "count": 1, "volume": 0.1, "batch_number": ""})
        unload_count += 1

    return tray_list, unload_count


def _parse_id_list(raw: str):
    text = (raw or "").strip()
    if not text:
        return []
    tokens = [t.strip() for t in text.replace("，", ",").split(",") if t.strip()]
    return tokens


def _calc_volume_m3_from_spec(spec: str, qty: int) -> float:
    parts = [p for p in str(spec).lower().replace(" ", "").split("x") if p]
    if len(parts) != 3:
        return 0.0
    a, b, c = [float(x) for x in parts]
    return (a * b * c * float(qty)) / 1_000_000_000.0


def _parse_spec_dims(spec: str):
    parts = [p for p in str(spec).lower().replace(" ", "").split("x") if p]
    if len(parts) != 3:
        raise ValueError(f"invalid spec: {spec}")
    return int(float(parts[0])), int(float(parts[1])), int(float(parts[2]))


def _normalize_spec(spec: str) -> str:
    return str(spec or "").strip().lower().replace(" ", "")


def _normalize_secondary_spec_text(spec: str) -> str:
    key = _normalize_spec(spec)
    parts = [p for p in key.split("x") if p]
    if len(parts) != 3:
        raise ValueError(f"invalid spec: {spec}")
    dims = []
    for p in parts:
        try:
            v = int(float(p))
        except Exception:
            raise ValueError(f"invalid spec: {spec}")
        if v <= 0:
            raise ValueError(f"invalid spec: {spec}")
        dims.append(str(v))
    return "x".join(dims)


def _secondary_rule_map(session):
    rules = {}
    for spec, base in SECONDARY_SPEC_PCS.items():
        key = _normalize_spec(spec)
        vals = [int(base)]
        for v in SECONDARY_SPEC_EXTRA_PCS.get(spec, []):
            try:
                iv = int(v)
            except Exception:
                continue
            if iv > 0 and iv not in vals:
                vals.append(iv)
        rules[key] = vals

    rows = session.query(FlowMetric).filter(FlowMetric.key.like(f"{SECONDARY_RULE_PREFIX}%")).all()
    for row in rows:
        k = str(row.key or "")
        if not k.startswith(SECONDARY_RULE_PREFIX):
            continue
        spec = _normalize_spec(k[len(SECONDARY_RULE_PREFIX):])
        if not spec:
            continue
        text = str(row.value or "")
        vals = []
        for token in text.replace("，", ",").split(","):
            token = token.strip()
            if not token:
                continue
            try:
                iv = int(float(token))
            except Exception:
                continue
            # 业务约束：pcs=1 属于异常历史脏值（会导致任意数量都命中该规格），读取时忽略。
            if iv > 1 and iv not in vals:
                vals.append(iv)
        if vals:
            existing = rules.get(spec, [])
            merged = list(existing)
            for iv in vals:
                if iv not in merged:
                    merged.append(iv)
            rules[spec] = merged
    return rules


def _save_secondary_rule(session, spec_key: str, values: list[int]):
    uniq = []
    for v in values:
        try:
            iv = int(v)
        except Exception:
            continue
        if iv > 0 and iv not in uniq:
            uniq.append(iv)
    if not uniq:
        return
    row = session.query(FlowMetric).filter_by(key=f"{SECONDARY_RULE_PREFIX}{spec_key}").first()
    text = ",".join(str(v) for v in sorted(uniq))
    if not row:
        row = FlowMetric(key=f"{SECONDARY_RULE_PREFIX}{spec_key}", value=text)
        session.add(row)
    else:
        row.value = text


def _register_secondary_rule(spec: str, pcs: int):
    spec_key = _normalize_secondary_spec_text(spec)
    pcs_val = _to_int(pcs, 0)
    # 业务约束：pcs=1 会破坏规格自动识别（任意数量都可被 1 整除），禁止写入。
    if pcs_val <= 1:
        raise ValueError("invalid pcs")
    session = Session()
    try:
        rules = _secondary_rule_map(session)
        vals = list(rules.get(spec_key, []))
        if pcs_val not in vals:
            vals.append(pcs_val)
            _save_secondary_rule(session, spec_key, vals)
            session.commit()
        return spec_key
    finally:
        session.close()


def _infer_spec_and_volume(product_id: str, pcs: int, spec_hint: str | None = None):
    if pcs <= 0:
        raise ValueError("invalid pcs")

    session = Session()
    try:
        rules = _secondary_rule_map(session)
    finally:
        session.close()

    if spec_hint:
        key = _normalize_secondary_spec_text(spec_hint)
        allowed = rules.get(key, [])
        if not allowed:
            raise ValueError(f"unsupported spec: {spec_hint}")
        if not any(pcs % fixed == 0 for fixed in allowed):
            allowed_text = "/".join(str(x) for x in allowed)
            raise ValueError(f"pcs must be a multiple of {allowed_text} for {key}")
        return key, _calc_volume_m3_from_spec(key, pcs)

    candidates = []
    for spec, allowed in rules.items():
        if any(pcs % fixed == 0 for fixed in allowed):
            candidates.append(spec)
    if not candidates:
        raise ValueError(f"pcs not matched to fixed rules: {pcs}")
    if len(candidates) > 1:
        raise ValueError(f"multiple specs for pcs {pcs}, please choose spec")
    spec = candidates[0]
    return spec, _calc_volume_m3_from_spec(spec, pcs)


def _read_inventory_data():
    rows = list_inventory_products(status="库存")
    product = {}
    for row in rows:
        product[row["product_id"]] = {
            "spec": row["spec"],
            "grade": row["grade"],
            "pcs": row["pcs"],
            "volume": row["volume"],
            "status": row["status"],
        }
    return {"product": product}


def _split_product_id(pid: str, item: dict):
    base = str(pid or "")
    grade = str(item.get("grade", "") or "").strip().upper()
    if "#" in base:
        b, g = base.split("#", 1)
        base = b.strip()
        if not grade:
            grade = g.strip().upper()
    return base, grade


def _parse_spec_3d(spec: str):
    parts = [p.strip() for p in str(spec or "").lower().replace(" ", "").split("x") if p.strip()]
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return parts[0], parts[1], ""
    if len(parts) == 1:
        return parts[0], "", ""
    return "", "", ""


def _parse_spec_dwl(spec: str):
    l, w, d = _parse_spec_3d(spec)
    return d, w, l


def _format_spec_dwl(spec: str) -> str:
    d, w, l = _parse_spec_dwl(spec)
    if d and w and l:
        return f"{d}x{w}x{l}"
    return str(spec or "")


def _collect_finished_product_rows():
    inv = _read_inventory_data()
    prod = inv.get("product", {}) if isinstance(inv.get("product"), dict) else {}
    rows = []
    for pid, item in prod.items():
        if not isinstance(item, dict):
            continue
        if item.get("status") != "库存":
            continue
        code_no, grade = _split_product_id(pid, item)
        spec = str(item.get("spec", "") or "")
        d_val, w_val, l_val = _parse_spec_dwl(spec)
        rows.append(
            {
                "product_id": pid,
                "编号": code_no,
                "D": d_val,
                "W": w_val,
                "L": l_val,
                "数量": int(_to_int(item.get("pcs"), 0)),
                "m³": float(_to_float(item.get("volume"), 0.0)),
                "等级": grade or str(item.get("grade", "") or "").strip().upper(),
                "重量(kg)": "",
                "规格": spec,
            }
        )
    rows.sort(key=lambda r: r.get("编号", ""))
    return rows


def _shipment_status_sort_value(status: str) -> int:
    order = {
        "待发车": 0,
        "去仰光途中": 1,
        "仰光仓已到": 2,
        "已从仰光出港": 3,
        "中国港口已到": 4,
        "异常": 5,
    }
    return order.get(str(status or "").strip(), 99)


def _next_shipment_no(shipping_data: dict) -> str:
    today = datetime.now().strftime("%Y%m%d")
    meta = shipping_data.get("meta", {}) if isinstance(shipping_data.get("meta"), dict) else {}
    last_date = str(meta.get("last_date", "") or "")
    last_seq = _to_int(meta.get("last_seq"), 0)
    seq = last_seq + 1 if last_date == today else 1
    shipping_data["meta"] = {"last_date": today, "last_seq": seq}
    return f"FH{today}{seq:03d}"


def _summarize_shipping_orders():
    data = get_shipping_data()
    shipments = data.get("shipments", []) if isinstance(data.get("shipments"), list) else []
    rows = []
    summary = {"去仰光途中": 0, "仰光仓已到": 0, "已从仰光出港": 0, "异常": 0}

    for item in shipments:
        if not isinstance(item, dict):
            continue
        status = str(item.get("status", "去仰光途中") or "去仰光途中")
        if status == "待发车":
            summary["去仰光途中"] = summary.get("去仰光途中", 0) + 1
        elif status in summary:
            summary[status] = summary.get(status, 0) + 1
        products = item.get("products", []) if isinstance(item.get("products"), list) else []
        total_pcs = sum(_to_int(p.get("pcs"), 0) for p in products if isinstance(p, dict))
        total_volume = sum(_to_float(p.get("volume"), 0.0) for p in products if isinstance(p, dict))
        rows.append(
            {
                "shipment_no": str(item.get("shipment_no", "") or ""),
                "customer": str(item.get("customer", "") or ""),
                "destination": str(item.get("destination", "") or ""),
                "vehicle_no": str(item.get("vehicle_no", "") or ""),
                "driver_name": str(item.get("driver_name", "") or ""),
                "tracking_no": str(item.get("tracking_no", "") or ""),
                "departure_at": str(item.get("departure_at", "") or ""),
                "eta_hours_to_yangon": _to_int(item.get("eta_hours_to_yangon"), 36),
                "yangon_arrived_at": str(item.get("yangon_arrived_at", "") or ""),
                "yangon_departed_at": str(item.get("yangon_departed_at", "") or ""),
                "china_port_arrived_at": str(item.get("china_port_arrived_at", "") or ""),
                "status": status,
                "product_count": len(products),
                "total_pcs": total_pcs,
                "total_volume": round(total_volume, 4),
                "updated_at": str(item.get("updated_at", "") or item.get("created_at", "") or ""),
                "created_at": str(item.get("created_at", "") or ""),
                "remark": str(item.get("remark", "") or ""),
                "products": products,
            }
        )

    rows.sort(key=lambda r: (_shipment_status_sort_value(r.get("status")), r.get("shipment_no", "")), reverse=False)
    return {"summary": summary, "rows": rows}


def _filter_finished_product_rows(grade: str = "", d: str = "", w: str = "", l: str = "", pcs: str = "", m3: str = ""):
    rows = _collect_finished_product_rows()
    grade = str(grade or "").strip().upper()
    d = str(d or "").strip()
    w = str(w or "").strip()
    l = str(l or "").strip()
    pcs = str(pcs or "").strip()
    m3 = str(m3 or "").strip()

    out = []
    for row in rows:
        if grade and str(row.get("等级", "")).upper() != grade:
            continue
        if d and str(row.get("D", "")) != d:
            continue
        if w and str(row.get("W", "")) != w:
            continue
        if l and str(row.get("L", "")) != l:
            continue
        if pcs and str(row.get("数量", "")) != pcs:
            continue
        if m3 and f"{float(row.get('m³', 0.0)):.4f}" != m3:
            continue
        out.append(row)
    return out


def _build_label_sheet(workbook: Workbook, rows: list, logo_path: Path):
    ws = workbook.active
    ws.title = "Labels-1"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    label_h = 9
    label_w = 8
    pages = max(1, (len(rows) + 7) // 8)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def ensure_sheet(idx: int):
        if idx == 0:
            return ws
        return workbook.create_sheet(f"Labels-{idx+1}")

    for page_idx in range(pages):
        sheet = ensure_sheet(page_idx)
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT

        for col in range(1, label_w * 2 + 1):
            sheet.column_dimensions[chr(64 + col)].width = 12
        for row_idx in range(1, label_h * 4 + 1):
            sheet.row_dimensions[row_idx].height = 21

        chunk = rows[page_idx * 8 : (page_idx + 1) * 8]
        for i in range(8):
            r_block = i // 2
            c_block = i % 2
            r0 = r_block * label_h + 1
            c0 = c_block * label_w + 1
            c1 = c0 + label_w - 1
            c2 = c0 + 2
            c3 = c0 + 3

            for rr in range(r0, r0 + 5):
                for cc in range(c0, c1 + 1):
                    cell = sheet.cell(rr, cc)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", horizontal="left")
                    cell.font = Font(size=10)

            sheet.merge_cells(start_row=r0, start_column=c3, end_row=r0, end_column=c1)
            sheet.cell(r0, c3).value = "DATE: ____________"
            sheet.cell(r0, c3).font = Font(size=10, bold=True)

            sheet.merge_cells(start_row=r0 + 1, start_column=c0, end_row=r0 + 1, end_column=c1)
            sheet.merge_cells(start_row=r0 + 2, start_column=c0, end_row=r0 + 2, end_column=c1)
            sheet.merge_cells(start_row=r0 + 3, start_column=c0, end_row=r0 + 3, end_column=c1)
            sheet.merge_cells(start_row=r0 + 4, start_column=c0, end_row=r0 + 4, end_column=c1)

            if i < len(chunk):
                item = chunk[i]
                code = item.get("编号", "")
                spec = item.get("规格", "")
                grade = str(item.get("等级", "") or item.get("grade", "") or "").strip().upper()
                qty = item.get("数量", 0)
                cbm = item.get("m³", 0.0)
                sheet.cell(r0 + 1, c0).value = f"CODE NO: {code}"
                sheet.cell(r0 + 2, c0).value = f"SIZE: {spec}    GRADE: {grade}"
                sheet.cell(r0 + 3, c0).value = f"PCS: {qty}    CBM: {cbm:.4f}    KG: ______"
            else:
                sheet.cell(r0 + 1, c0).value = "CODE NO: ____________"
                sheet.cell(r0 + 2, c0).value = "SIZE: ____________    GRADE: ____"
                sheet.cell(r0 + 3, c0).value = "PCS: ____    CBM: ____    KG: ____"
            sheet.cell(r0 + 4, c0).value = "QC SIGN: ____________"

            if logo_path.exists():
                try:
                    img = XLImage(str(logo_path))
                    img.width = 70
                    img.height = 20
                    sheet.add_image(img, sheet.cell(r0, c0).coordinate)
                except Exception:
                    sheet.cell(r0, c0).value = "LOGO"
            else:
                sheet.cell(r0, c0).value = "LOGO"


def _ensure_template_logos(ws, logo_path: Path):
    if not logo_path.exists():
        return

    def _col_width_to_pixels(width):
        width = float(width or 8.43)
        return int(width * 7 + 5)

    def _row_height_to_pixels(height):
        height = float(height or 15)
        return int(height * 96 / 72)

    try:
        ws._images = []
        logo_slots = ["A1", "D1", "A7", "D7", "A13", "D13", "A19", "D19"]
        for cell in logo_slots:
            col_letter = cell[0]
            row_idx = int(cell[1:])
            cell_w = _col_width_to_pixels(ws.column_dimensions[col_letter].width)
            cell_h = _row_height_to_pixels(ws.row_dimensions[row_idx].height)
            img_w = min(137, max(100, cell_w - 10))
            img_h = int(img_w * 39 / 137)
            if img_h > cell_h - 6:
                img_h = max(24, cell_h - 6)
                img_w = int(img_h * 137 / 39)
            offset_x = max(0, (cell_w - img_w) // 2)
            offset_y = max(0, (cell_h - img_h) // 2)

            img = XLImage(str(logo_path))
            img.width = img_w
            img.height = img_h
            marker = AnchorMarker(
                col=ord(col_letter) - ord("A"),
                colOff=pixels_to_EMU(offset_x),
                row=row_idx - 1,
                rowOff=pixels_to_EMU(offset_y),
            )
            img.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(pixels_to_EMU(img_w), pixels_to_EMU(img_h)),
            )
            ws.add_image(img)
    except ImportError:
        # openpyxl requires Pillow for adding images at runtime.
        # Keep export working even if the runtime environment lacks Pillow.
        return


def _underlined_chars(value: str) -> str:
    raw = str(value or "")
    if raw:
        text = f" {raw} "
    else:
        text = "  "
    if not text:
        return ""
    return "".join(f"{ch}\u0332" for ch in text)


def _fill_template_label_sheet(ws, rows: list):
    slots = [
        ("A", "B", 1),
        ("D", "E", 1),
        ("A", "B", 7),
        ("D", "E", 7),
        ("A", "B", 13),
        ("D", "E", 13),
        ("A", "B", 19),
        ("D", "E", 19),
    ]
    for idx, (logo_col, text_col, start_row) in enumerate(slots):
        item = rows[idx] if idx < len(rows) else {}
        code = str(item.get("编号", "") or "")
        spec = _format_spec_dwl(str(item.get("规格", "") or ""))
        grade = str(item.get("等级", "") or item.get("grade", "") or "").strip().upper()
        qty = item.get("数量", 0) or 0
        cbm = float(item.get("m³", 0.0) or 0.0)
        date_text = datetime.now().strftime("%Y-%m-%d")

        top_cell = ws[f"{text_col}{start_row}"]
        code_cell = ws[f"{logo_col}{start_row + 1}"]
        size_cell = ws[f"{logo_col}{start_row + 2}"]
        grade_cell = ws[f"{text_col}{start_row + 2}"]
        pcs_cell = ws[f"{logo_col}{start_row + 3}"]
        qc_cell = ws[f"{logo_col}{start_row + 4}"]

        top_cell.value = f"DATE:{_underlined_chars(date_text)}"
        code_cell.value = f"CODE NO:{_underlined_chars(code)}"
        size_cell.value = f"SIZE:{_underlined_chars(spec)}"
        grade_cell.value = f"GRADE:{_underlined_chars(grade)}"
        pcs_cell.value = f"PCS:{_underlined_chars(str(qty))}  CBM:{_underlined_chars(f'{cbm:.4f}')}  KG:{_underlined_chars('')}"
        qc_cell.value = f"QC SIGN:{_underlined_chars('')}"

        top_cell.font = top_cell.font.copy(sz=13)
        code_cell.font = code_cell.font.copy(sz=13)
        size_cell.font = size_cell.font.copy(sz=13)
        grade_cell.font = grade_cell.font.copy(sz=13)
        pcs_cell.font = pcs_cell.font.copy(sz=13)
        qc_cell.font = qc_cell.font.copy(sz=13)
        top_cell.alignment = top_cell.alignment.copy(horizontal="left", vertical="center")
        code_cell.alignment = code_cell.alignment.copy(horizontal="left", vertical="center")
        size_cell.alignment = size_cell.alignment.copy(horizontal="left", vertical="center")
        grade_cell.alignment = grade_cell.alignment.copy(horizontal="left", vertical="center")
        pcs_cell.alignment = pcs_cell.alignment.copy(horizontal="left", vertical="center")
        qc_cell.alignment = qc_cell.alignment.copy(horizontal="left", vertical="center")


def _build_label_workbook_from_template(template_path: Path, rows: list, logo_path: Path):
    wb = load_workbook(template_path)
    base_ws = wb[wb.sheetnames[0]]
    per_sheet = 8
    pages = max(1, (len(rows) + per_sheet - 1) // per_sheet)

    for page_idx in range(pages):
        if page_idx == 0:
            ws = base_ws
        else:
            ws = wb.copy_worksheet(base_ws)
            ws.title = f"{base_ws.title}-{page_idx + 1}"
        _ensure_template_logos(ws, logo_path)
        chunk = rows[page_idx * per_sheet : (page_idx + 1) * per_sheet]
        _fill_template_label_sheet(ws, chunk)

    return wb


def _resolve_logo_path() -> Path:
    primary = BASE / "static" / "logo.png"
    if primary.exists():
        return primary
    return BASE / "web" / "static" / "logo.png"
